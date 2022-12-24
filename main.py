from cProfile import label
from fileinput import filename
from fractions import Fraction
from re import X
import tkinter as tk
# tkinter._test()
from tkinter import BOTTOM, TOP, Y, Button, Frame, Label, filedialog as fd
from tkinter import ttk
from tkinter.font import BOLD
from tkinter.messagebox import showinfo
#Regx
from distutils.filelist import findall
import openpyxl
from openpyxl.styles import colors, Color, PatternFill
import re
import pandas as pd



# filename = fd.askopenfilename()

#create the root window
root = tk.Tk()
root.title('Negative Keywords Sorter')
root.resizable(False, False)
root.geometry('600x300')


def select_file():
    f= []
    filetypes = (
        ('text files', '*.xlsx'),
        ('All files', '*.*')
    )
    filename = fd.askopenfilename(
        title='Open a file',
        initialdir='/',
        filetypes=filetypes)

    showinfo(
        title='Selected File',
        message=filename
    )
    f.append(filename)
    print(f[0])
    check(f)
    # sort_key_neg(f)
    # return filename

    # print("a")
    
    # return f.append(filename)

# def filename():
#     return select_file()
    
    
def sort_key_neg():
    # if f:
    #     print(f[0])
    # for i in filename:
        # path = i
        # print("aaaaa",i)
    path = "C:\\Users\\User\\Documents\\N_LIMS_S_USA_May21.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    
    for x in range(1,1519,1):
    
        cell_obj = sheet_obj.cell(x, column = 1)
        y= re.findall(r"srl", cell_obj.value)
        w= re.findall(r"\Aa", cell_obj.value)
        a = re.findall(r"chem", cell_obj.value)
        d = re.findall(r"phar", cell_obj.value)

        if y or a or d:
            print(cell_obj.value)
            cell_obj.fill =  PatternFill("solid", start_color="5cb800")        
     
        if w:
            # print(cell_obj.value)
            cell_obj.fill =  PatternFill("solid", start_color="5cb800")
        
    wb_obj.save("N_LIMS_S_USA_May21_2.xlsx")
    

#Title of Software
heading = Label(text= "Sort Negative Keywords", bg="green", fg="white", font=("comicsans", 18, BOLD))
heading.pack(fill="x")

dev = Label(text= "Wellcome Shrihari ", font=("comicsans", 16))
dev.pack()

footer = Frame(root, bg="white", borderwidth=6)
footer.pack(side = BOTTOM, fill="x")
Name= Label(footer, text="Develped For CrelioHealth @ 2022")
Name.pack()

select_file1= Frame(root, bg="white", borderwidth=6)
select_file1.pack(side = TOP, fill="x", pady=10, padx=20)
file = Label(select_file1, text="File Name")
file.pack()


#
button_frame = Frame(root, borderwidth=6)
button_frame.pack(side = TOP, fill="x")
# open button
open_button = ttk.Button(
    button_frame,
    text='Open a File',
    command=select_file
)

open_button.pack(expand=True)

sort = ttk.Button(button_frame, text= 'Sort', command=sort_key_neg)
sort.pack(expand=True, pady=5)

# file = Label(text="File name:")
# file.pack()

# a = "2222"

def check(f):
    if f:
        filename = Label( select_file1, bg="white", text = f[0])
        filename.pack()
        
    else:
        print("empty")
        return 2
# filename = Label(text = check())
# filename.pack()

# if f:
#     filename = Label(text = f[0])
#     filename.pack()
    
# Sorting function 


root.mainloop()